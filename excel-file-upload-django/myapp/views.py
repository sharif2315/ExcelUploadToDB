from django.shortcuts import render
import openpyxl

from .models import *

def home(request):
    employees = Employee.objects.all()
    context = { 'employees': employees }
    return render(request, 'myapp/home.html', context)


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        # print(active_sheet)

        # reading a cell
        # print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        # removing header from excel dataset
        excel_data.remove(excel_data[0])

        exceluploadData = []
        for row in excel_data:
            newdict = {}
            newdict = {}
            newdict['name'] = row[0]
            newdict['age'] = row[1]
            newdict['gender'] = row[2]
            exceluploadData.append(newdict)

        print(excel_data)

        # Posting Data to database
        # for row in excel_data:
        #     _, created = Employee.objects.update_or_create(
        #     name = row[0],
        #     age = int(row[1]),
        #     gender = row[2])

        context = { "excel_data": excel_data, 'exceluploadData': exceluploadData }

        return render(request, 'myapp/index.html', context)







        #
